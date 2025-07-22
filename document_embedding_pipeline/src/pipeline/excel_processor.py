"""
Excel processing pipeline that plugs into the existing state-machine.
Keeps the rich parsing logic of the GitHub ExcelParser, including
complex table finding.
"""
from __future__ import annotations
import json
import logging
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from services.llm_service import LLMService
from utils.logging_setup import setup_logging
from pipeline.base_processor import BaseProcessor
from data_models import Headers, Summary, SerializedRows, ExcelDocumentPayload, Header, Title, TableSheet, ContentSheet
from utils.table_parsing import ParsedTableSheet, ParsedContentSheet, ParsedSheetFactory

logger = logging.getLogger(__name__)

class ExcelProcessor(BaseProcessor):
    """
    Excel equivalent of WordProcessor.
    Implements:
        process(file_path: Path) -> Dict[str, Any]
    """

    def __init__(
        self,
        temp_folder : Path,
    ):
        """
        Initializes the ExcelProcessor.

        Args:
            temp_folder : Path
        """
        self.llm = LLMService()
        self.parsed_sheet_factory = ParsedSheetFactory()

    async def process(self, file_path: Path) -> Dict[str, Any]:
        """
        Full async pipeline for a single Excel file.
        Returns the same dict structure that WordProcessor returns so the
        orchestrator can treat both uniformly, but enriched with metadata
        and detailed cell info from ExcelParser.
        """
        logger.info("Starting Excel processing: %s", file_path.name)

        workbook = load_workbook(file_path, data_only=True)
        try:
            sheets : List[ParsedContentSheet | ParsedTableSheet]= []
            for sheet_name in workbook.sheetnames:
                sheet = self.parsed_sheet_factory.create_correct_sheet_object(workbook[sheet_name], sheet_name)
                sheets.append(sheet)

            content_sheets = [sheet for sheet in sheets if isinstance(sheet, ParsedContentSheet)]
            for sheet in sheets:       
                if isinstance(sheet, ParsedTableSheet):
                    logger.info(f"Parsing table sheet {sheet.name}")
                    await sheet.extract_tables(content_sheets, self.llm)

            title = await self.extract_title_from_doc(file_path, content_sheets)
            
            final_document = self.format_data_for_storage(title, file_path, sheets)
            return final_document

        finally:
            workbook.close()

    async def extract_title_from_doc(self, file_path : Path, content_sheets : List[ParsedContentSheet]):
        title = None
        content_for_title = ""
        if content_sheets: # If it is not empty
            for sheet in content_sheets:
                content_for_title += f"Sheet called '{sheet.name}' : \n"
                content_for_title += sheet.export_blocks_as_markdown()

        content_for_title = content_for_title[:min(len(content_for_title), 4000)] # Title is rarely found very deep into the doc, so we limit the size of the search space
        if len(content_for_title) < 50: # If the content to search in is that small its because there is no content sheet, or content sheets are empty
            logger.warning("Falling back to file name since context for title extraction is too small")
            title = file_path.name
        else:
            logger.info("Extracting title")
            title = await self._extract_title(content_for_title)
        if title is None:
            logger.warning("LLM found no title, falling back to the filename")
            title = file_path.name
        return title

    async def _extract_title(
        self,
        content : str
    ) -> List[str]:
        """Extract the document title from the first two pages"""

        prompt = (
            "Given the sheets below, find what name the full document has. Only return the verbatim doc title.\n"
            f"{content}\n"
            "Output in JSON"
        )
        resp: Title = await self.llm.aget_structured_response(
            prompt, response_model=Title
        )
        return resp.title
    
    def format_data_for_storage(self, title: str, file_path:str, sheets : List[ParsedContentSheet | ParsedTableSheet]):
        formatted_sheets : List[TableSheet, ContentSheet]= []
        for idx, sheet in enumerate(sheets):
            if isinstance(sheet, ParsedTableSheet):
                for table in sheet.tables:
                    table_sheet = TableSheet(sheet_name=f"{sheet.name}{idx}", table_headers=table.headers, table_summary=table.summary, serialized_rows=table.data_rows)
                    formatted_sheets.append(table_sheet)
            else:
                content_sheet = ContentSheet(sheet_name=sheet.name, content=sheet.export_blocks_as_markdown(), summary=sheet.summary)
                formatted_sheets.append(content_sheet)

        excel_doc = ExcelDocumentPayload(title=title, file_path=str(file_path), sheets=formatted_sheets)

        excel_doc_rep = ""
        for sheet in excel_doc.sheets:
            if isinstance(sheet, TableSheet):
                excel_doc_rep += sheet.table_summary
        logger.info(excel_doc_rep)
        return excel_doc