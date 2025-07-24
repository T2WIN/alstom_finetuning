from typing import Dict, List
import openpyxl

def classify_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> str:
    """
    Classifies the worksheet as either 'table' or 'content'.

    Returns
    -------
    str
        'table'  – if the sheet contains at least one block with > 100 filled cells  
        'content' – otherwise
    """
    blocks: List[Dict] = _find_all_blocks(sheet)

    # Check the “table” criterion
    if any(b["non_empty"] > 80 for b in blocks):
        return "table"
    content = ""
    for idx, block in enumerate(blocks):
                        content += f"Block {idx} :"
                        content += _tiny_markdown_table(sheet, block)
                        content += "\n"
    print(content)
    return "content"


# ------------------------------------------------------------------
# The helpers below are unchanged
# ------------------------------------------------------------------
def _tiny_markdown_table(ws, block, rows=8, cols=6):
    """
    Returns a compact markdown table (no header, ≤ rows×cols cells)
    for the top-left corner of the block.
    """
    r1, r2 = block["r1"], block["r2"]
    c1, c2 = block["c1"], block["c2"]
    lines = []
    for r in range(r1, min(r1 + rows, r2 + 1)):
        cells = [
            str(ws.cell(row=r, column=c).value or "")
            for c in range(c1, min(c1 + cols, c2 + 1))
        ]
        lines.append("| " + " | ".join(cells) + " |")
    if r2 - r1 + 1 > rows:
        lines[-1] += "\n⋮"
    if c2 - c1 + 1 > cols:
        lines = [ln + " …" for ln in lines]
    return "\n".join(lines) if lines else "(empty)"

def _find_all_blocks(ws):
    """
    Quick, lightweight scan to return every rectangular block of
    non-empty cells (including single-cell blocks).
    Returns list[dict] with r1, r2, c1, c2, rows, cols, non_empty.
    """
    visited = [[False] * (ws.max_column + 1) for _ in range(ws.max_row + 1)]
    blocks = []

    def _dfs(r, c):
        stack = [(r, c)]
        min_r = max_r = r
        min_c = max_c = c
        filled = 0
        while stack:
            x, y = stack.pop()
            if x < 1 or y < 1 or x > ws.max_row or y > ws.max_column:
                continue
            if visited[x][y] or ws.cell(row=x, column=y).value is None:
                continue
            visited[x][y] = True
            filled += 1
            min_r, max_r = min(min_r, x), max(max_r, x)
            min_c, max_c = min(min_c, y), max(max_c, y)
            for dx, dy in ((0, 1), (1, 0), (0, -1), (-1, 0)):
                stack.append((x + dx, y + dy))
        return {
            "r1": min_r, "r2": max_r, "c1": min_c, "c2": max_c,
            "rows": max_r - min_r + 1,
            "cols": max_c - min_c + 1,
            "non_empty": filled,
        }

    for row in range(1, min(ws.max_row + 1, 500)):
        for col in range(1, min(ws.max_column + 1, 100)):
            if not visited[row][col] and ws.cell(row=row, column=col).value is not None:
                blocks.append(_dfs(row, col))
    return blocks

wb = openpyxl.load_workbook("/home/grand/alstom_finetuning/data/SSPHA projets/TRB_BATTERY MAKE_S03_002_SSPHA Battery next gen_A1.xlsx")
sheet = wb["Front Page (template)"]
print(classify_sheet(sheet))  # → "table" or "content"