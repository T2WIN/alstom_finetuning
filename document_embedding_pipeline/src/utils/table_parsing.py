import dataclasses
from typing import List, Optional, Dict, Any
from openpyxl.worksheet.worksheet import Worksheet 
from data_models import Summary, Title
from services.llm_service import LLMService
import logging
import json


logger = logging.getLogger(__name__)

@dataclasses.dataclass
class BlockInfo:
    """A data structure to hold information about a contiguous block of cells."""
    r1: int  # Top-most row index
    c1: int  # Left-most column index
    r2: int  # Bottom-most row index
    c2: int  # Right-most column index
    rows: int # Total number of rows in the block's bounding box
    cols: int # Total number of columns in the block's bounding box
    non_empty_cells: int # Count of non-empty cells within the block

@dataclasses.dataclass
class CellData:
    """Represents enriched data for a single cell."""
    value: Any
    header: Optional[str]
    row: int
    column: int
    # Potentially more fields like data_type, style, etc.

@dataclasses.dataclass
class TableBoundary:
    """Defines the rectangular coordinates of a table."""
    start_row: int
    start_col: int
    end_row: int
    end_col: int

@dataclasses.dataclass
class ExtractedTable:
    """Contains all information about an extracted table."""
    headers: List[str]
    data_rows: List[List[CellData]]
    location: TableBoundary

@dataclasses.dataclass
class FinalTable:
    """Contains all information that I need from an Excel table."""
    headers: List[str]
    data_rows: List[dict]
    summary : str




class ParsedSheetFactory:

    def create_correct_sheet_object(self, worksheet : Worksheet, name : str):
        classification, blocks = self.classify_worksheet(worksheet)
        if classification == "table":
            return ParsedTableSheet(worksheet, name)
        else:
            return ParsedContentSheet(worksheet, name, blocks)

    def classify_worksheet(self, worksheet : Worksheet):
        """
        Classifies the worksheet as either 'table' or 'content'.
        We limit the number of rows and columns to load to 100 to avoid loading entire tables when they get very large.

        Returns
        -------
        str
            'table'  – if the sheet contains at least one block with more filled cells than the threshold 
            'content' – otherwise
        """
        blocks: List[BlockInfo] = self._find_contiguous_blocks(worksheet, scan_limit=100)
        if any(b.non_empty_cells > 80 for b in blocks):
            return "table", blocks

        return "content", blocks
    
    def _find_contiguous_blocks(self, worksheet: Worksheet, scan_limit: Optional[int] = None) -> List[BlockInfo]:
        """
        Identifies all separate, contiguous blocks of non-empty cells in a worksheet.

        This function scans the worksheet and uses a Depth-First Search (DFS) algorithm
        to find groups of adjacent cells that contain data.

        Args:
            ws: The worksheet object to analyze (e.g., from openpyxl).
            scan_limit: An optional integer to limit the scan area to a square of
                        (scan_limit x scan_limit) cells for performance on very large sheets.
                        If None, the entire sheet is scanned.

        Returns:
            A list of BlockInfo objects, where each object describes the
            bounding box and cell count of a contiguous block.
        """
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # If a scan limit is provided, use it to cap the iteration range.
        scan_rows = min(max_row, scan_limit) if scan_limit else max_row
        scan_cols = min(max_col, scan_limit) if scan_limit else max_col

        # A 2D grid to keep track of cells that have already been visited
        # as part of a block, preventing redundant checks.
        visited = [[False] * (max_col + 1) for _ in range(max_row + 1)]
        blocks = []

        def _find_single_block_dfs(start_row: int, start_col: int) -> BlockInfo:
            """
            Performs a Depth-First Search (DFS) starting from a single cell
            to find the full extent of its contiguous block.
            """
            # A stack is used for an iterative DFS. It starts with the first found cell.
            stack = [(start_row, start_col)]
            
            # Initialize the bounding box of the block to the starting cell's coordinates.
            min_r, max_r = start_row, start_row
            min_c, max_c = start_col, start_col
            
            filled_cell_count = 0

            while stack:
                curr_row, curr_col = stack.pop()

                # --- Boundary and Condition Checks ---
                # 1. Ensure the cell is within the worksheet's dimensions.
                is_out_of_bounds = (curr_row < 1 or curr_col < 1 or
                                    curr_row > max_row or curr_col > max_col)
                if is_out_of_bounds:
                    continue
                
                # 2. Skip if the cell was already visited or is empty.
                if visited[curr_row][curr_col] or worksheet.cell(row=curr_row, column=curr_col).value is None:
                    continue

                # --- Process the Cell ---
                # Mark the current cell as visited.
                visited[curr_row][curr_col] = True
                filled_cell_count += 1
                
                # Expand the bounding box to include the current cell.
                min_r = min(min_r, curr_row)
                max_r = max(max_r, curr_row)
                min_c = min(min_c, curr_col)
                max_c = max(max_c, curr_col)

                # Add all four adjacent neighbors (up, down, left, right) to the stack for future processing.
                for dr, dc in [(0, 1), (0, -1), (1, 0), (-1, 0)]:
                    stack.append((curr_row + dr, curr_col + dc))
            
            return BlockInfo(
                r1=min_r, c1=min_c, r2=max_r, c2=max_c,
                rows=max_r - min_r + 1,
                cols=max_c - min_c + 1,
                non_empty_cells=filled_cell_count
            )

        # Iterate through each cell in the worksheet (or within the scan_limit).
        for r in range(1, scan_rows + 1):
            for c in range(1, scan_cols + 1):
                # If a cell contains data and has not yet been visited, it's the start of a new block.
                if not visited[r][c] and worksheet.cell(row=r, column=c).value is not None:
                    # Launch a DFS to find all connected cells in this new block.
                    block_info = _find_single_block_dfs(r, c)
                    blocks.append(block_info)
        
        return blocks
        
class ParsedContentSheet:

    def __init__(self, worksheet : Worksheet, worksheet_name: str, blocks : List[BlockInfo]):
        self.ws = worksheet
        self.name = worksheet_name
        self.summary = "summary"
        self.content = None
        self.blocks = blocks

    def export_blocks_as_markdown(self):
        """
        Returns markdown representation of the contiguous blocks of the sheet.
        """
        if self.content:
            return self.content
        
        markdown_representation = ""
        for idx, block in enumerate(self.blocks):
            lines = []
            for r in range(block.r1, block.r2+1):
                cells = [
                    str(self.ws.cell(row=r, column=c).value or "")
                    for c in range(block.c1, block.c2 + 1)
                ]
                lines.append("| " + " | ".join(cells) + " |")
            block_representation = f"Block {idx} :" + "\n".join(lines) if lines else "(empty)"
            markdown_representation += block_representation

        self.content = markdown_representation
        return markdown_representation

    

class ParsedTableSheet:

    def __init__(self, worksheet : Worksheet, worksheet_name: str):
        self.ws = worksheet
        self.name = worksheet_name
        self.tables : List[FinalTable] = []

    @staticmethod
    def _enrich_cell(cell, header: Optional[str], row: int, col: int) -> CellData:
        """
        A helper to convert a raw cell into a structured CellData object.
        This can be expanded to include more complex parsing or type detection.
        """
        return CellData(
            value=cell.value,
            header=header,
            row=row,
            column=col
        )
    
    async def extract_tables(
        self, content_sheets : List[ParsedContentSheet], llm : LLMService
    ) -> List[FinalTable]:
        """
        Parse a list of found tables, generate summaries + row texts.
        """
        tables = self.find_all_tables()
        final_tables = []
        for tbl in tables:
            summary = await self._summarise_table(tbl, content_sheets, llm)
            rows = self.extract_rows_from_table(tbl)
            headers = tbl.headers
            final_tables.append(FinalTable(headers, rows, summary))         

        self.tables = final_tables

    def _determine_table_boundary(self, start_row: int, start_col: int) -> TableBoundary:
        """
        Calculates the rectangular boundary of a table starting from a given cell.

        It works by:
        1. Finding the right edge (end_col) by locating the first empty column.
        2. Finding the bottom edge (end_row) by locating the first empty row within that width.
        """
        # 1. Determine the table's width (find end_col)
        # We expand rightwards, column by column, until we hit a column with no data.
        end_col = start_col
        for c in range(start_col, self.ws.max_column + 2): # +2 to check one past the end
            # Check if this entire column (from the start_row downwards) is empty.
            has_data_in_col = any(self.ws.cell(row=r, column=c).value is not None for r in range(start_row, self.ws.max_row + 1))
            if has_data_in_col:
                end_col = c # This column has data, so it's part of the table.
            else:
                break # First empty column found, table width is determined.

        # 2. Determine the table's height (find end_row)
        # With a fixed width, we expand downwards, row by row, until we hit an empty row.
        end_row = start_row
        for r in range(start_row, self.ws.max_row + 2): # +2 to check one past the end
            # Check if this row has any data within the table's determined columns.
            has_data_in_row = any(self.ws.cell(row=r, column=c).value is not None for c in range(start_col, end_col + 1))
            if has_data_in_row:
                end_row = r # This row has data, so it's part of the table.
            else:
                break # First empty row found, table height is determined.

        return TableBoundary(start_row, start_col, end_row, end_col)
    
    def find_all_tables(self) -> List[ExtractedTable]:
        """
        Scans a worksheet to find and extract all distinct, structured tables.

        A table is assumed to start with a header row and is identified by finding
        a cell with a string value that hasn't been included in a previously found table.

        Args:
            sheet: The worksheet object to analyze.

        Returns:
            A list of ExtractedTable objects, each representing a table found.
        """
        tables = []
        # A set to keep track of cells that are already part of a table,
        # preventing them from being treated as the start of a new table.
        processed_cells = set()

        # Iterate through every cell to find potential starting points for tables.
        for r_idx in range(1, self.ws.max_row + 1):
            for c_idx in range(1, self.ws.max_column + 1):
                # --- Condition to start a new table search ---
                # 1. The cell must not have been processed yet.
                # 2. The cell should contain a string, as we assume tables start with a text header.
                if (r_idx, c_idx) in processed_cells:
                    continue
                
                cell = self.ws.cell(row=r_idx, column=c_idx)
                if not (cell.value and isinstance(cell.value, str)):
                    continue

                # --- Found a potential table, now extract it ---
                boundary = self._determine_table_boundary(r_idx, c_idx)

                # Extract header information from the first row of the boundary
                header_cells = [
                    self._enrich_cell(self.ws.cell(row=boundary.start_row, column=c), None, boundary.start_row, c)
                    for c in range(boundary.start_col, boundary.end_col + 1)
                ]
                
                # If the header row is completely empty, it's not a valid table.
                if not any(cell.value is not None for cell in header_cells):
                    continue

                headers = [str(cell.value or "") for cell in header_cells]

                # Extract data from all subsequent rows within the boundary
                data_rows = []
                for r in range(boundary.start_row + 1, boundary.end_row + 1):
                    row_data = []
                    for c in range(boundary.start_col, boundary.end_col + 1):
                        header_text = headers[c - boundary.start_col]
                        cell_obj = self.ws.cell(row=r, column=c)
                        cell_data = self._enrich_cell(cell_obj, header_text, r, c)
                        row_data.append(cell_data)
                    data_rows.append(row_data)
                
                # If the table has no data rows, we can choose to ignore it.
                if not data_rows:
                    continue

                # Create the final table object
                table = ExtractedTable(
                    headers=headers,
                    data_rows=data_rows,
                    location=boundary
                )
                tables.append(table)

                # Mark all cells within the extracted table's boundary as processed.
                for r in range(boundary.start_row, boundary.end_row + 1):
                    for c in range(boundary.start_col, boundary.end_col + 1):
                        processed_cells.add((r, c))

        return tables
    
    async def _summarise_table(
        self, table : ExtractedTable, content_sheets: List[ParsedContentSheet], llm : LLMService
    ) -> str:
        """Generate a concise natural-language summary of the table."""
        formatted_context = ""
        for idx, content_sheet in enumerate(content_sheets):
            formatted_context += f"Sheet {idx} called {content_sheet.name} :\n"
            formatted_context += content_sheet.export_blocks_as_markdown()
            formatted_context += "\n"

        sample = self.extract_rows_from_table(table)[4:7]

        if len(formatted_context) > 20:
            prompt = (
                "Summarise the following table in 1-2 sentences:\n"
                f"Headers: {table.headers}\n"
                f"Sample rows: {json.dumps(sample, ensure_ascii=False, indent=2)}"
                f"Here is context gathered from other sheets that might help you understand : {formatted_context}"
            )
        else:
           prompt = (
                "Summarise the following table in 1-2 sentences:\n"
                f"Headers: {table.headers}\n"
                f"Sample rows: {json.dumps(sample, ensure_ascii=False, indent=2)}"
            ) 
        resp: Summary = await llm.aget_structured_response(
            prompt, max_tokens=512, response_model=Summary
        )
        return resp.summary
    
    @staticmethod
    def extract_rows_from_table(table: ExtractedTable) -> List[dict]:
        """Extracts a dict representation of a row"""
        rows = []
        for row in table.data_rows:
            row_dict = {}
            for cell in row:
                row_dict[cell.header] = str(cell.value)
            rows.append(row_dict)
        return rows
    

#     async def _rewrite_headers(self, table: Dict[str, Any]) -> List[str]:
#         """Use LLM to clean / rewrite column headers."""
#         sample_rows = [[c["value"] for c in r] for r in table["data"][4:7]]
#         prompt = (
#             "You are a data-analysis expert.\n"
#             f"Original headers: {table['headers']}\n"
#             f"First 3 rows: {sample_rows}\n\n"
#             """Example : {
#   "headers": [
#     {
#       "name": "Employee ID",
#       "description": "A unique identifier assigned to each employee."
#     },
#     {
#       "name": "Full Name",
#       "description": "The first and last name of the employee."
#     },
#     {
#       "name": "Department",
#       "description": "The department within the company where the employee works."
#     },
#     {
#       "name": "Start Date",
#       "description": "The date the employee began their employment with the company."
#     }
#   ]
# }"""
#             "Return only the new headers, same order, cleaned with a short description based on the values seen in the rows."
#         )
#         headers: Headers = await self.llm.aget_structured_response(
#             prompt, max_tokens=1024, response_model=Headers
#         )
#         return headers